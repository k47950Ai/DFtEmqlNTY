# 代码生成时间: 2025-09-20 00:56:28
import cherrypy
def read_excel_file(file_path):
    """
    读取Excel文件的内容
    :param file_path: Excel文件路径
    :return: Excel文件内容
    """
    try:
        import openpyxl
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))
        return data
    except Exception as e:
        raise Exception(f"Failed to read Excel file: {e}")
def generate_excel(data, file_name):
    """
    生成Excel文件
    :param data: 要写入的数据
    :param file_name: 生成的Excel文件名
    :return: None
    """
    try:
        import openpyxl
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for row in data:
            sheet.append(row)
        workbook.save(file_name)
    except Exception as e:
        raise Exception(f"Failed to generate Excel file: {e}")class ExcelGenerator:
    """
    Excel表格自动生成器
    """
    @cherrypy.expose
    def index(self):
        """
        首页
        """
        return "Welcome to Excel Generator"
    @cherrypy.expose
    def upload(self, file=None):
        """
        上传Excel文件
        :param file: 上传的文件
        :return: 上传结果
        """
        if file:
            file_path = file.filename
            data = read_excel_file(file_path)
            return f"File uploaded successfully. Data: {data}"
        else:
            return "No file uploaded"    @cherrypy.expose
    def generate(self, data=None, file_name=None):
        """
        生成Excel文件
        :param data: 要写入的数据
        :param file_name: 生成的Excel文件名
        :return: 生成结果
        """
        if data and file_name:
            generate_excel(data, file_name)
            return f"Excel file generated successfully. File name: {file_name}"
        else:
            return "No data or file name provided"if __name__ == '__main__':
    conf = {
        "global": {"server.socket_host": '127.0.0.1',
                   "server.socket_port": 8080,
                   "tools.sessions.on": True,
                   "tools.sessions.timeout": 60 * 60 * 12},
        '/': {'tools.sessions.on': True}}
    cherrypy.quickstart(ExcelGenerator(), '/', conf)